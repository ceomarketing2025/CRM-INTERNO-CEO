import json

from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.core.exceptions import PermissionDenied
from apps.audit.models import ActivityLog
from apps.audit.services import log_activity
from apps.core.decorators import role_required
from apps.operations.models import WebProductionSheet
from apps.projects.models import Project
from apps.projects.selectors import can_access_project
from .forms import ProjectQuestionnaireCreateForm
from .models import Answer, ProjectQuestionnaire, QuestionnaireTemplate
from .models.choices import AnswerState, QuestionnaireStatus
from .models.seo_status import SEOProjectStatus, SEOUrlStatus
from django.urls import reverse

from openpyxl import load_workbook


from apps.reminders.models import Reminder


from .services import (
    WEBSITE_TEMPLATE_CODE,
    save_key_answer,
    website_answer_payload,
    website_questionnaire_progress,
)


def _clean(value):
    return (value or "").strip()


def _yes_no(value):
    return value if value in {"yes", "no"} else ""
def _seo_parse_date(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    if not text:
        return None

    formats = (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y",
        "%d/%m/%Y",
    )

    for date_format in formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            ).date()
        except ValueError:
            continue

    return None


def _seo_parse_decimal(value):
    if value in {
        None,
        "",
        "-",
    }:
        return None

    try:
        return Decimal(
            str(value)
            .replace(",", "")
            .strip()
        )
    except (
        InvalidOperation,
        ValueError,
        TypeError,
    ):
        return None


def _seo_normalize_status(value):
    text = (
        str(value or "")
        .strip()
        .lower()
    )

    normalized = (
        text
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )

    if normalized in {
        "indexed",
        "indexada",
        "indexado",
        "si",
        "yes",
    }:
        return "indexed"

    if normalized in {
        "not indexed",
        "not_indexed",
        "no indexada",
        "no indexado",
        "no",
    }:
        return "not_indexed"

    if normalized in {
        "pending",
        "pendiente",
        "en proceso",
    }:
        return "pending"

    return "unknown"


def _seo_normalize_alert(value):
    text = (
        str(value or "")
        .strip()
        .lower()
    )

    normalized = (
        text
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )

    if normalized in {
        "ok",
        "sin alerta",
        "bien",
    }:
        return "ok"

    if normalized in {
        "attention",
        "atencion",
    }:
        return "attention"

    if normalized in {
        "critical",
        "critica",
        "critico",
    }:
        return "critical"

    return "monitor"

def _indexed_rows(request, prefix, fields):
    """Read dynamic rows sent as prefix_<field>[] while preserving row position."""
    columns = {field: request.POST.getlist(f"{prefix}_{field}") for field in fields}
    count = max([len(values) for values in columns.values()] or [0])
    rows = []
    for index in range(count):
        row = {field: _clean(columns[field][index]) if index < len(columns[field]) else "" for field in fields}
        if any(row.values()):
            rows.append(row)
    return rows


def _status_for(condition, negative=False):
    return bool(condition), bool(negative and condition)


def _website_save(questionnaire, request):
    user = request.user

    company_description = _clean(request.POST.get("company_description"))
    save_key_answer(
        questionnaire=questionnaire, key="company_description", user=user,
        value_text=company_description, complete=bool(company_description),
    )

    # ==========================================
    # LOGICA DEL NEGOCIO
    # ==========================================

    customer_type = _clean(
        request.POST.get("customer_type")
    )

    business_sells = _clean(
        request.POST.get("business_sells")
    )

    business_audience = _clean(
        request.POST.get("business_audience")
    )

    business_where = _clean(
        request.POST.get("business_where")
    )

    business_priority_sale = _clean(
        request.POST.get("business_priority_sale")
    )

    visitor_action = _clean(
        request.POST.get("visitor_action")
    )

    business_logic_complete = all([
        customer_type,
        business_sells,
        business_audience,
        business_where,
        business_priority_sale,
        visitor_action,
    ])

    business_logic_lines = []

    if customer_type:
        business_logic_lines.append(
            f"Tipo de cliente: {customer_type}"
        )

    if business_sells:
        business_logic_lines.append(
            f"Qué vende: {business_sells}"
        )

    if business_audience:
        business_logic_lines.append(
            f"A quién le vende: {business_audience}"
        )

    if business_where:
        business_logic_lines.append(
            f"Dónde vende: {business_where}"
        )

    if business_priority_sale:
        business_logic_lines.append(
            f"Qué quiere vender más: {business_priority_sale}"
        )

    if visitor_action:
        business_logic_lines.append(
            f"Acción principal: {visitor_action}"
        )

    save_key_answer(
        questionnaire=questionnaire,
        key="business_logic",
        user=user,
        value_text="\n".join(
            business_logic_lines
        ),
        value_json={
            "customer_type": customer_type,
            "sells": business_sells,
            "audience": business_audience,
            "where": business_where,
            "priority_sale": business_priority_sale,
            "visitor_action": visitor_action,
        },
        complete=business_logic_complete,
    )


    # ==========================================
    # COTIZACIONES
    # ==========================================

    free_estimates = _clean(
        request.POST.get("free_estimates")
    )

    quote_channels = [
        _clean(value)
        for value in request.POST.getlist(
            "quote_channel"
        )
        if _clean(value)
    ]

    quote_notes = _clean(
        request.POST.get("quote_notes")
    )

    quotes_complete = (
        bool(free_estimates) and
        bool(quote_channels)
    )

    quote_lines = []

    if free_estimates:
        quote_lines.append(
            f"Estimados gratuitos: {free_estimates}"
        )

    if quote_channels:
        quote_lines.append(
            "Canales: " +
            ", ".join(quote_channels)
        )

    if quote_notes:
        quote_lines.append(
            f"Notas: {quote_notes}"
        )

    save_key_answer(
        questionnaire=questionnaire,
        key="quotes",
        user=user,
        value_text="\n".join(
            quote_lines
        ),
        value_json={
            "free_estimates": free_estimates,
            "channels": quote_channels,
            "notes": quote_notes,
        },
        complete=quotes_complete,
    )




    # ==========================================
    # FORMULARIOS
    # ==========================================

    form_fields = [
        _clean(value)
        for value in request.POST.getlist(
            "form_fields"
        )
        if _clean(value)
    ]

    form_receiver_email = _clean(
        request.POST.get(
            "form_receiver_email"
        )
    )

    form_receiver_whatsapp = _clean(
        request.POST.get(
            "form_receiver_whatsapp"
        )
    )

    form_confirmation_message = _clean(
        request.POST.get(
            "form_confirmation_message"
        )
    )

    form_redirect = _clean(
        request.POST.get(
            "form_redirect"
        )
    )

    forms_complete = (
        bool(form_fields) and
        bool(
            form_receiver_email or
            form_receiver_whatsapp
        )
    )

    forms_lines = []

    if form_fields:
        forms_lines.append(
            "Campos: " +
            ", ".join(form_fields)
        )

    if form_receiver_email:
        forms_lines.append(
            f"Email receptor: {form_receiver_email}"
        )

    if form_receiver_whatsapp:
        forms_lines.append(
            "WhatsApp receptor: " +
            form_receiver_whatsapp
        )

    if form_confirmation_message:
        forms_lines.append(
            "Confirmación: " +
            form_confirmation_message
        )

    if form_redirect:
        forms_lines.append(
            f"Redirección: {form_redirect}"
        )

    save_key_answer(
        questionnaire=questionnaire,
        key="website_forms",
        user=user,
        value_text="\n".join(
            forms_lines
        ),
        value_json={
            "fields": form_fields,
            "receiver_email":
                form_receiver_email,
            "receiver_whatsapp":
                form_receiver_whatsapp,
            "confirmation_message":
                form_confirmation_message,
            "redirect":
                form_redirect,
        },
        complete=forms_complete,
    )
    slogan_mode = request.POST.get("slogan_mode", "")
    slogan_text = _clean(request.POST.get("slogan_text"))
    slogan_complete = slogan_mode == "none" or (slogan_mode in {"existing", "create"} and bool(slogan_text))
    slogan_label = {
        "existing": f"Slogan actual: {slogan_text}",
        "create": f"Slogan creado/propuesto: {slogan_text}",
        "none": "No crear slogan",
    }.get(slogan_mode, slogan_text)
    save_key_answer(
        questionnaire=questionnaire, key="slogan", user=user,
        value_text=slogan_label, value_json={"mode": slogan_mode, "text": slogan_text},
        complete=slogan_complete, negative=slogan_mode == "none",
    )

    mission_has = _yes_no(request.POST.get("mission_has"))
    mission = _clean(request.POST.get("mission_text"))
    vision = _clean(request.POST.get("vision_text"))
    mission_origin = _clean(request.POST.get("mission_origin"))
    mission_clients = _clean(request.POST.get("mission_clients"))
    vision_future = _clean(request.POST.get("vision_future"))
    vision_legacy = _clean(request.POST.get("vision_legacy"))
    mission_complete = (
        mission_has == "yes" and bool(mission and vision)
    ) or (
        mission_has == "no" and all([mission_origin, mission_clients, vision_future, vision_legacy])
    )
    mission_summary = ""
    if mission_has == "yes":
        mission_summary = f"Misión: {mission}\nVisión: {vision}".strip()
    elif mission_has == "no":
        mission_summary = "Información recopilada para crear misión y visión"
    save_key_answer(
        questionnaire=questionnaire, key="has_mission_vision", user=user,
        value_text=mission_summary,
        value_json={
            "has": mission_has, "mission": mission, "vision": vision,
            "mission_origin": mission_origin, "mission_clients": mission_clients,
            "vision_future": vision_future, "vision_legacy": vision_legacy,
        }, complete=mission_complete, negative=False,
    )

        # ==========================================
    # ABOUT Y TEAM
    # ==========================================

    team_include = _yes_no(
        request.POST.get(
            "team_include"
        )
    )

    team_scope = _clean(
        request.POST.get(
            "team_scope"
        )
    )

    team_other = _clean(
        request.POST.get(
            "team_other"
        )
    )

    team_notes = _clean(
        request.POST.get(
            "team_notes"
        )
    )


    allowed_team_scopes = {
        "owner_photo",
        "team_photo",
        "group_photos",
        "no_photo",
        "other",
    }


    if team_scope not in allowed_team_scopes:
        team_scope = ""


    if team_include == "no":

        team_scope = ""
        team_other = ""
        team_notes = ""

        team_complete = True

        team_summary = (
            "No incluir contenido de About / Team"
        )


    elif team_include == "yes":

        if team_scope == "other":

            team_complete = bool(
                team_other
            )

        else:

            team_complete = (
                team_scope
                in allowed_team_scopes
                and team_scope != "other"
            )


        team_labels = {
            "owner_photo":
                "Foto del dueño",

            "team_photo":
                "Foto con el team",

            "group_photos":
                "Fotos por grupos",

            "no_photo":
                "Sin foto",

            "other":
                team_other or "Otro",
        }


        team_summary = (
            team_labels.get(
                team_scope,
                ""
            )
        )


        if team_notes:

            team_summary = (
                f"{team_summary}\n"
                f"Notas: {team_notes}"
            ).strip()


    else:

        team_complete = False
        team_summary = ""


    save_key_answer(
        questionnaire=questionnaire,
        key="show_team",
        user=user,
        value_text=team_summary,
        value_json={
            "include": team_include,
            "scope": team_scope,
            "other": team_other,
            "notes": team_notes,
        },
        complete=team_complete,
        negative=(
            team_include == "no"
        ),
    )
    service_names = [
        _clean(name)
        for name in request.POST.getlist(
            "service_name"
        )
    ]

    service_names = [
        name
        for name in service_names
        if name
    ]


    primary_raw = request.POST.get(
        "service_primary",
        ""
    )

    try:
        primary_index = int(
            primary_raw
        )
    except (
        TypeError,
        ValueError,
    ):
        primary_index = None


    business_type = _clean(
        request.POST.get(
            "services_business_type"
        )
    )

    business_type_other = _clean(
        request.POST.get(
            "business_type_other"
        )
    )


    structure_raw = request.POST.get(
        "service_structure_json",
        ""
    )

    try:
        categories = (
            json.loads(structure_raw)
            if structure_raw
            else []
        )
    except (
        json.JSONDecodeError,
        TypeError,
    ):
        categories = []


    excluded_services = [
        _clean(name)
        for name in request.POST.getlist(
            "excluded_service_name"
        )
    ]

    excluded_services = [
        name
        for name in excluded_services
        if name
    ]


    services = []

    for index, name in enumerate(
        service_names
    ):

        services.append({
            "name": name,
            "primary": (
                index == primary_index
            ),
            "excluded": False,
            "order": index,
        })


    for name in excluded_services:

        services.append({
            "name": name,
            "primary": False,
            "excluded": True,
            "order": len(services),
        })


    found_primary = False

    for item in services:

        if (
            item["primary"]
            and not item["excluded"]
            and not found_primary
        ):
            found_primary = True

        else:
            item["primary"] = False


    service_lines = []

    for item in services:

        line = item["name"]

        if item["primary"]:
            line += " (PRINCIPAL)"

        if item["excluded"]:
            line += " (DEJAR FUERA)"

        service_lines.append(
            line
        )


    save_key_answer(
        questionnaire=questionnaire,
        key="main_services",
        user=user,
        value_text="\n".join(
            service_lines
        ),
        value_json={
            "business_type": business_type,
            "business_type_other": business_type_other,
            "items": services,
            "categories": categories,
            "excluded_services": excluded_services,
        },
        complete=bool(
            business_type
            and service_names
            and found_primary
            and (
                business_type != "other"
                or business_type_other
            )
        ),
    )

    area_types = request.POST.getlist("area_type")
    area_names = request.POST.getlist("area_name")
    areas = []
    for index, raw_name in enumerate(area_names):
        name = _clean(raw_name)
        if not name:
            continue
        area_type = area_types[index] if index < len(area_types) and area_types[index] in {"state", "county", "city"} else "city"
        areas.append({"type": area_type, "name": name})
    area_labels = {"state": "Estado", "county": "Condado", "city": "Ciudad"}
    save_key_answer(
        questionnaire=questionnaire, key="service_areas", user=user,
        value_text="\n".join(f"{area_labels.get(item['type'], 'Área')}: {item['name']}" for item in areas),
        value_json={"items": areas}, complete=bool(areas),
    )

    coverage_miles = _clean(request.POST.get("coverage_miles"))
    coverage_complete = coverage_miles.isdigit() and int(coverage_miles) >= 0
    save_key_answer(
        questionnaire=questionnaire, key="coverage", user=user,
        value_text=f"{coverage_miles} millas alrededor" if coverage_miles else "",
        value_json={"miles": coverage_miles}, complete=coverage_complete,
    )



    # ==========================================
    # HORARIO Y ATENCION 24/7
    # ==========================================

    is_24_7 = _yes_no(
        request.POST.get(
            "is_24_7"
        )
    )

    start_time = _clean(
        request.POST.get(
            "hours_start"
        )
    )

    start_period = (
        request.POST.get(
            "hours_start_period",
            "AM",
        )
        if request.POST.get(
            "hours_start_period"
        ) in {
            "AM",
            "PM",
        }
        else "AM"
    )

    end_time = _clean(
        request.POST.get(
            "hours_end"
        )
    )

    end_period = (
        request.POST.get(
            "hours_end_period",
            "PM",
        )
        if request.POST.get(
            "hours_end_period"
        ) in {
            "AM",
            "PM",
        }
        else "PM"
    )


    hours_complete = bool(
        start_time
        and end_time
    )


    hours_text = (
        f"{start_time} {start_period} - "
        f"{end_time} {end_period}"
        if hours_complete
        else ""
    )


    save_key_answer(
        questionnaire=questionnaire,
        key="business_hours",
        user=user,
        value_text=hours_text,
        value_json={
            "start": start_time,
            "start_period": start_period,
            "end": end_time,
            "end_period": end_period,
        },
        complete=hours_complete,
    )


    save_key_answer(
        questionnaire=questionnaire,
        key="is_24_7",
        user=user,
        value_text=(
            "Sí"
            if is_24_7 == "yes"
            else (
                "No"
                if is_24_7 == "no"
                else ""
            )
        ),
        value_json={
            "answer": is_24_7
        },
        complete=(
            is_24_7
            in {
                "yes",
                "no",
            }
        ),
        negative=(
            is_24_7 == "no"
        ),
    )

    # ==========================================
    # TELEFONOS
    # ==========================================

    phone_numbers = (
        request.POST.getlist(
            "phone_number"
        )
    )

    phone_primary = (
        request.POST.get(
            "phone_primary",
            "",
        )
    )

    phone_whatsapp_indexes = set(
        request.POST.getlist(
            "phone_whatsapp"
        )
    )


    phones = []


    for index, raw_number in enumerate(
        phone_numbers
    ):

        number = _clean(
            raw_number
        )

        if not number:
            continue


        phones.append({
            "number": number,

            "primary": (
                str(index)
                == str(
                    phone_primary
                )
            ),

            "whatsapp": (
                str(index)
                in phone_whatsapp_indexes
            ),
        })


    if (
        phones
        and not any(
            item["primary"]
            for item in phones
        )
    ):

        phones[0][
            "primary"
        ] = True


    phone_lines = []


    for item in phones:

        tags = []


        if item["primary"]:

            tags.append(
                "Principal"
            )


        if item["whatsapp"]:

            tags.append(
                "WhatsApp"
            )


        suffix = (
            f" ({', '.join(tags)})"
            if tags
            else ""
        )


        phone_lines.append(
            f"{item['number']}"
            f"{suffix}"
        )


    save_key_answer(
        questionnaire=questionnaire,
        key="contact_numbers",
        user=user,
        value_text="\n".join(
            phone_lines
        ),
        value_json={
            "numbers": phones
        },
        complete=bool(
            phones
        ),
    )
    social_has = _yes_no(request.POST.get("social_has"))
    social_keys = ["facebook", "instagram", "tiktok", "youtube", "twitter", "yelp", "linkedin", "nextdoor"]
    networks = {}
    for key in social_keys:
        if request.POST.get(f"social_enabled_{key}") == "1":
            networks[key] = _clean(request.POST.get(f"social_url_{key}"))
    custom_names = request.POST.getlist("social_custom_name")
    custom_urls = request.POST.getlist("social_custom_url")
    custom_social = []
    for index in range(max(len(custom_names), len(custom_urls))):
        name = _clean(custom_names[index]) if index < len(custom_names) else ""
        url = _clean(custom_urls[index]) if index < len(custom_urls) else ""
        if name or url:
            custom_social.append({"name": name, "url": url})
    has_social_link = any(networks.values()) or any(item.get("url") for item in custom_social)
    social_complete = social_has == "no" or (social_has == "yes" and has_social_link)
    social_labels = {
        "facebook": "Facebook", "instagram": "Instagram", "tiktok": "TikTok", "youtube": "YouTube",
        "twitter": "X / Twitter", "yelp": "Yelp", "linkedin": "LinkedIn", "nextdoor": "Nextdoor",
    }
    social_lines = [f"{social_labels[key]}: {url}" for key, url in networks.items() if url]
    social_lines += [f"{item['name'] or 'Otra red'}: {item['url']}" for item in custom_social if item.get("url")]
    social_text = "No tiene redes sociales" if social_has == "no" else "\n".join(social_lines)
    save_key_answer(
        questionnaire=questionnaire, key="has_social", user=user,
        value_text=social_text, value_json={"has": social_has, "networks": networks, "custom": custom_social},
        complete=social_complete, negative=social_has == "no",
    )

 




    cert_has = _yes_no(request.POST.get("certifications_has"))
    cert_detail = _clean(request.POST.get("certifications_detail"))
    cert_complete = cert_has == "no" or (cert_has == "yes" and bool(cert_detail))
    save_key_answer(
        questionnaire=questionnaire, key="certifications", user=user,
        value_text=cert_detail if cert_has == "yes" else ("No tiene certificaciones" if cert_has == "no" else ""),
        value_json={"has": cert_has, "detail": cert_detail}, complete=cert_complete, negative=cert_has == "no",
    )

    warranty_has = _yes_no(request.POST.get("warranties_has"))
    warranty_detail = _clean(request.POST.get("warranties_detail"))
    warranty_complete = warranty_has == "no" or (warranty_has == "yes" and bool(warranty_detail))
    save_key_answer(
        questionnaire=questionnaire, key="warranties", user=user,
        value_text=warranty_detail if warranty_has == "yes" else ("No ofrece garantía" if warranty_has == "no" else ""),
        value_json={"has": warranty_has, "detail": warranty_detail}, complete=warranty_complete, negative=warranty_has == "no",
    )

    experience_years = _clean(request.POST.get("experience_years"))
    experience_complete = experience_years.isdigit()
    save_key_answer(
        questionnaire=questionnaire, key="experience_years", user=user,
        value_text=experience_years, value_json={"years": experience_years}, complete=experience_complete,
    )

    

    progress = website_questionnaire_progress(
        questionnaire
)

    questionnaire.status = (
        QuestionnaireStatus.COMPLETE
        if progress["percent"] == 100
        else QuestionnaireStatus.IN_PROGRESS
    )

    questionnaire.save(
        update_fields=[
            "status",
            "updated_at",
        ]
    )

    return progress


def _website_fill_context(questionnaire):
    payload = website_answer_payload(questionnaire)
    progress = website_questionnaire_progress(questionnaire)
    return {
        "questionnaire": questionnaire,
        "data": payload,
        "progress": progress,
        "social_networks": [
            ("facebook", "Facebook"), ("instagram", "Instagram"), ("tiktok", "TikTok"),
            ("youtube", "YouTube"), ("twitter", "X / Twitter"), ("yelp", "Yelp"),
            ("linkedin", "LinkedIn"), ("nextdoor", "Nextdoor"),
        ],
    }


@role_required("developer")
def development_dashboard(request):
    projects = Project.objects.select_related("client", "purchased_plan__plan").filter(project_type__in=["website", "software", "seo"])
    if not request.user.is_manager:
        projects = projects.filter(assignments__user=request.user, assignments__area="development").distinct()

    rows = []
    for project in projects:
        questionnaire = ProjectQuestionnaire.objects.filter(project=project).select_related("template").order_by("id").first()
        technical_progress = website_questionnaire_progress(questionnaire) if questionnaire and questionnaire.template.code == WEBSITE_TEMPLATE_CODE else {
            "complete": 1 if questionnaire and questionnaire.status == QuestionnaireStatus.COMPLETE else 0,
            "total": 1 if questionnaire else 0,
            "percent": 100 if questionnaire and questionnaire.status == QuestionnaireStatus.COMPLETE else 0,
        }
        sheet = WebProductionSheet.objects.filter(project=project).prefetch_related("pages", "counties__services", "cities").first()
        last_log = ActivityLog.objects.filter(metadata__project_id=project.pk).select_related("user").first()
        if sheet:
            all_rows = list(sheet.pages.all()) + list(sheet.counties.all()) + [s for c in sheet.counties.all() for s in c.services.all()] + list(sheet.cities.all())
            total = len(all_rows)
            complete = sum(1 for row in all_rows if row.state == "complete")
            production_percent = round((complete / total) * 100) if total else 0
        else:
            total = 0
            complete = 0
            production_percent = 0
        overall_parts = [technical_progress["percent"]]
        if sheet:
            overall_parts.append(production_percent)
        overall_percent = round(sum(overall_parts) / len(overall_parts)) if overall_parts else 0
        rows.append({
            "project": project,
            "questionnaire": questionnaire,
            "questionnaire_done": bool(questionnaire and questionnaire.status == QuestionnaireStatus.COMPLETE),
            "technical_progress": technical_progress,
            "sheet": sheet,
            "production_total": total,
            "production_complete": complete,
            "production_percent": production_percent,
            "overall_percent": overall_percent,
            "last_log": last_log,
        })
    in_progress_count = sum(
        1 for row in rows
        if 0 < row["overall_percent"] < 100
    )
    completed_count = sum(
        1 for row in rows
        if row["overall_percent"] == 100
    )

    reminders_count = Reminder.objects.filter(
        assigned_to=request.user,
        status="pending",
    ).count()
    return render(
        request,
        "questionnaires/development_dashboard.html",
        {
            "rows": rows,
            "in_progress_count": in_progress_count,
            "completed_count": completed_count,
            "reminders_count": reminders_count,
        },
    )

@role_required("developer")
def seo_status(request):

    scope = request.GET.get("scope", "all")

    projects = (
        Project.objects
        .select_related(
            "client",
            "purchased_plan__plan",
        )
        .filter(
            project_type__in=[
                "website",
                "seo",
            ]
        )
    )

    if scope == "mine":
        projects = projects.filter(
            assignments__user=request.user
        ).distinct()

    rows = []

    for project in projects:

        seo_project = (
            SEOProjectStatus.objects
            .filter(project=project)
            .prefetch_related("urls")
            .first()
        )

        if seo_project:

            total_urls = seo_project.total_urls
            indexed_urls = seo_project.indexed_urls
            indexing_percent = seo_project.indexing_percent
            top_10_count = seo_project.top_10_count
            improved_count = seo_project.improved_count
            worsened_count = seo_project.worsened_count

        else:

            total_urls = 0
            indexed_urls = 0
            indexing_percent = 0
            top_10_count = 0
            improved_count = 0
            worsened_count = 0


        rows.append({
            "project": project,
            "seo_project": seo_project,
            "total_urls": total_urls,
            "indexed_urls": indexed_urls,
            "indexing_percent": indexing_percent,
            "top_10_count": top_10_count,
            "improved_count": improved_count,
            "worsened_count": worsened_count,
        })


    return render(
        request,
        "questionnaires/seo_status.html",
        {
            "rows": rows,
            "scope": scope,
        },
    )

@role_required("developer")
def seo_status_import(request):

    if request.method == "GET":

        preview = request.session.get(
            "seo_import_preview",
            [],
        )

        return render(
            request,
            "questionnaires/seo_status_import.html",
            {
                "preview": preview,
                "projects": (
                    Project.objects
                    .select_related(
                        "client"
                    )
                    .filter(
                        project_type__in=[
                            "website",
                            "seo",
                        ]
                    )
                    .order_by(
                        "client__business_name"
                    )
                ),
            },
        )

    action = _clean(
        request.POST.get(
            "action"
        )
    )

    if action == "save_import":

        preview = request.session.get(
            "seo_import_preview",
            [],
        )

        if not preview:
            messages.error(
                request,
                "No existe una importación pendiente para guardar.",
            )

            return redirect(
                "questionnaires:seo_status_import"
            )

        saved_projects = 0
        saved_urls = 0

        with transaction.atomic():

            for index, sheet_data in enumerate(
                preview
            ):

                project_id = _clean(
                    request.POST.get(
                        f"project_id_{index}"
                    )
                )

                if not project_id:
                    continue

                project = get_object_or_404(
                    Project,
                    pk=project_id,
                    project_type__in=[
                        "website",
                        "seo",
                    ],
                )

                seo_project, created = (
                    SEOProjectStatus.objects
                    .get_or_create(
                        project=project,
                        defaults={
                            "created_by":
                                request.user,
                        },
                    )
                )

                seo_project.sitemap_url = (
                    sheet_data.get(
                        "sitemap_url",
                        "",
                    )
                    or ""
                )

                seo_project.last_sitemap_read = (
                    _seo_parse_date(
                        sheet_data.get(
                            "last_sitemap_read"
                        )
                    )
                )

                if (
                    created
                    and not seo_project.created_by
                ):
                    seo_project.created_by = (
                        request.user
                    )

                seo_project.save()

                saved_projects += 1

                rows = (
                    sheet_data.get(
                        "rows",
                        []
                    )
                    or []
                )

                for order, row in enumerate(
                    rows
                ):

                    url = _clean(
                        row.get(
                            "url"
                        )
                    )

                    if not url:
                        continue

                    page_name = (
                        _clean(
                            row.get(
                                "page_name"
                            )
                        )
                        or url
                    )

                    defaults = {
                        "entry_date":
                            _seo_parse_date(
                                row.get(
                                    "entry_date"
                                )
                            ),

                        "page_name":
                            page_name,

                        "objective":
                            _clean(
                                row.get(
                                    "objective"
                                )
                            ),

                        "status":
                            _seo_normalize_status(
                                row.get(
                                    "status"
                                )
                            ),

                        "current_position":
                            _seo_parse_decimal(
                                row.get(
                                    "current_position"
                                )
                            ),

                        "previous_position":
                            _seo_parse_decimal(
                                row.get(
                                    "previous_position"
                                )
                            ),

                        "last_update":
                            _seo_parse_date(
                                row.get(
                                    "last_update"
                                )
                            ),

                        "last_sitemap_read":
                            _seo_parse_date(
                                row.get(
                                    "last_sitemap_read"
                                )
                            ),

                        "pending_action":
                            _clean(
                                row.get(
                                    "pending_action"
                                )
                            ),

                        "observations":
                            _clean(
                                row.get(
                                    "observations"
                                )
                            ),

                        "alert":
                            _seo_normalize_alert(
                                row.get(
                                    "alert"
                                )
                            ),

                        "order":
                            order,

                        "updated_by":
                            request.user,
                    }

                    seo_url, url_created = (
                        SEOUrlStatus.objects
                        .update_or_create(
                            seo_project=
                                seo_project,
                            url=url,
                            defaults=defaults,
                        )
                    )

                    if (
                        url_created
                        and not seo_url.created_by
                    ):
                        seo_url.created_by = (
                            request.user
                        )

                        seo_url.save(
                            update_fields=[
                                "created_by",
                            ]
                        )

                    saved_urls += 1

        if not saved_projects:

            messages.error(
                request,
                "Selecciona al menos un proyecto para guardar la importación.",
            )

            return redirect(
                "questionnaires:seo_status_import"
            )

        request.session.pop(
            "seo_import_preview",
            None,
        )

        messages.success(
            request,
            (
                f"Importación guardada: "
                f"{saved_projects} proyecto(s) y "
                f"{saved_urls} URL(s)."
            ),
        )

        return redirect(
            "questionnaires:seo_status"
        )

    uploaded_file = request.FILES.get(
        "seo_file"
    )

    if not uploaded_file:
        messages.error(
            request,
            "Selecciona un archivo Excel."
        )
        return redirect(
            "questionnaires:seo_status_import"
        )

    if not uploaded_file.name.lower().endswith(".xlsx"):
        messages.error(
            request,
            "El archivo debe ser formato .xlsx."
        )
        return redirect(
            "questionnaires:seo_status_import"
        )

    try:
        workbook = load_workbook(
            uploaded_file,
            data_only=True,
            read_only=True,
        )

    except Exception:
        messages.error(
            request,
            "No se pudo leer el archivo Excel."
        )
        return redirect(
            "questionnaires:seo_status_import"
        )

    preview = []

    for sheet_name in workbook.sheetnames:

        sheet = workbook[sheet_name]

        project_name = sheet["B3"].value
        domain = sheet["F3"].value
        sitemap_url = sheet["N3"].value
        last_cut = sheet["F4"].value
        last_sitemap_read = sheet["N4"].value

        rows = []

        for row_number in range(21, 121):

            page_name = sheet.cell(
                row=row_number,
                column=3,
            ).value

            url = sheet.cell(
                row=row_number,
                column=4,
            ).value

            if not page_name and not url:
                continue

            entry_date = sheet.cell(
                row=row_number,
                column=2,
            ).value

            objective = sheet.cell(
                row=row_number,
                column=6,
            ).value

            status = sheet.cell(
                row=row_number,
                column=7,
            ).value

            current_position = sheet.cell(
                row=row_number,
                column=9,
            ).value

            last_update = sheet.cell(
                row=row_number,
                column=11,
            ).value

            row_sitemap_read = sheet.cell(
                row=row_number,
                column=15,
            ).value

            pending_action = sheet.cell(
                row=row_number,
                column=19,
            ).value

            observations = sheet.cell(
                row=row_number,
                column=21,
            ).value

            alert = sheet.cell(
                row=row_number,
                column=22,
            ).value

            previous_position = sheet.cell(
                row=row_number,
                column=23,
            ).value


            rows.append({
                "row_number": row_number,
                "entry_date": entry_date,
                "page_name": page_name or "",
                "url": url or "",
                "objective": objective or "",
                "status": status or "",
                "current_position": current_position,
                "last_update": last_update,
                "last_sitemap_read": row_sitemap_read,
                "pending_action": pending_action or "",
                "observations": observations or "",
                "alert": alert or "",
                "previous_position": previous_position,
            })


        preview.append({
            "sheet_name": sheet_name,
            "project_name": project_name or sheet_name,
            "domain": domain or "",
            "sitemap_url": sitemap_url or "",
            "last_cut": last_cut,
            "last_sitemap_read": last_sitemap_read,
            "total_urls": len(rows),
            "rows": rows,
        })


    preview = json.loads(
        json.dumps(
            preview,
            default=str,
        )
    )

    request.session[
        "seo_import_preview"
    ] = preview


    return render(
        request,
        "questionnaires/seo_status_import.html",
        {
            "preview": preview,
            "projects": Project.objects
                .select_related("client")
                .filter(
                    project_type__in=[
                        "website",
                        "seo",
                    ]
                )
                .order_by(
                    "client__business_name"
                ),
        },
    )
@role_required("developer")
def seo_status_detail(
    request,
    project_pk,
):
    project = get_object_or_404(
        Project.objects.select_related(
            "client",
            "purchased_plan__plan",
        ),
        pk=project_pk,
        project_type__in=[
            "website",
            "seo",
        ],
    )

    if not can_access_project(
        request.user,
        project,
    ):
        raise PermissionDenied(
            "Este proyecto no está asignado a Desarrollo."
        )

    seo_project = get_object_or_404(
        SEOProjectStatus.objects
        .select_related(
            "project",
            "project__client",
            "created_by",
        )
        .prefetch_related(
            "urls__created_by",
            "urls__updated_by",
        ),
        project=project,
    )

    if request.method == "POST":

        action = _clean(
            request.POST.get("action")
        )

        if action == "save_url":

            url_id = _clean(
                request.POST.get("url_id")
            )

            seo_url = get_object_or_404(
                SEOUrlStatus,
                pk=url_id,
                seo_project=seo_project,
            )

            seo_url.page_name = _clean(
                request.POST.get("page_name")
            )

            seo_url.url = _clean(
                request.POST.get("url")
            )

            seo_url.entry_date = _seo_parse_date(
                request.POST.get("entry_date")
            )

            seo_url.objective = _clean(
                request.POST.get("objective")
            )

            seo_url.status = _seo_normalize_status(
                request.POST.get("status")
            )

            seo_url.current_position = _seo_parse_decimal(
                request.POST.get("current_position")
            )

            seo_url.previous_position = _seo_parse_decimal(
                request.POST.get("previous_position")
            )

            seo_url.last_update = _seo_parse_date(
                request.POST.get("last_update")
            )

            seo_url.last_sitemap_read = _seo_parse_date(
                request.POST.get("last_sitemap_read")
            )

            seo_url.pending_action = _clean(
                request.POST.get("pending_action")
            )

            seo_url.observations = _clean(
                request.POST.get("observations")
            )

            seo_url.alert = _seo_normalize_alert(
                request.POST.get("alert")
            )

            seo_url.updated_by = request.user

            seo_url.save()

            messages.success(
                request,
                f"{seo_url.page_name} actualizada correctamente.",
            )

            return redirect(
                "questionnaires:seo_status_detail",
                project_pk=project.pk,
            )

    urls = list(
        seo_project.urls.all()
        .select_related(
            "created_by",
            "updated_by",
        )
        .order_by(
            "order",
            "id",
        )
    )

    total_urls = len(urls)

    indexed_urls = sum(
        1
        for item in urls
        if item.status == "indexed"
    )

    not_indexed_urls = sum(
        1
        for item in urls
        if item.status == "not_indexed"
    )

    pending_urls = sum(
        1
        for item in urls
        if item.status == "pending"
    )

    unknown_urls = sum(
        1
        for item in urls
        if item.status == "unknown"
    )

    top_10_count = sum(
        1
        for item in urls
        if (
            item.current_position is not None
            and Decimal("1") <= item.current_position <= Decimal("10")
        )
    )

    top_20_count = sum(
        1
        for item in urls
        if (
            item.current_position is not None
            and Decimal("11") <= item.current_position <= Decimal("20")
        )
    )

    over_20_count = sum(
        1
        for item in urls
        if (
            item.current_position is not None
            and item.current_position > Decimal("20")
        )
    )

    improved_count = sum(
        1
        for item in urls
        if item.evolution_type == "improved"
    )

    worsened_count = sum(
        1
        for item in urls
        if item.evolution_type == "worsened"
    )

    stable_count = sum(
        1
        for item in urls
        if item.evolution_type == "stable"
    )

    indexing_percent = (
        round(
            (
                indexed_urls
                / total_urls
            )
            * 100
        )
        if total_urls
        else 0
    )

    return render(
        request,
        "questionnaires/seo_status_detail.html",
        {
            "project": project,
            "seo_project": seo_project,
            "urls": urls,

            "total_urls": total_urls,
            "indexed_urls": indexed_urls,
            "not_indexed_urls": not_indexed_urls,
            "pending_urls": pending_urls,
            "unknown_urls": unknown_urls,

            "indexing_percent": indexing_percent,

            "top_10_count": top_10_count,
            "top_20_count": top_20_count,
            "over_20_count": over_20_count,

            "improved_count": improved_count,
            "worsened_count": worsened_count,
            "stable_count": stable_count,
        },
    )

@role_required("developer")
def create_for_project(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk)
    if not can_access_project(request.user, project):
        raise PermissionDenied("Este proyecto no está asignado a Desarrollo.")

    if request.method == "GET" and request.GET.get("auto") == "1":
        template = QuestionnaireTemplate.objects.filter(project_type=project.project_type, is_active=True).order_by("id").first()
        if not template and project.project_type == "seo":
            template = QuestionnaireTemplate.objects.filter(project_type="website", is_active=True).order_by("id").first()
        if template:
            obj, _ = ProjectQuestionnaire.objects.get_or_create(
                project=project,
                template=template,
                defaults={"created_by": request.user, "status": QuestionnaireStatus.DRAFT},
            )
            log_activity(request.user, "questionnaires", "create", obj)
            return redirect("questionnaires:fill", pk=obj.pk)

    form = ProjectQuestionnaireCreateForm(request.POST or None, project=project)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.project = project
        obj.created_by = request.user
        if ProjectQuestionnaire.objects.filter(project=project, template=obj.template).exists():
            messages.warning(request, "Ese cuestionario ya existe en el proyecto.")
            return redirect("projects:detail", pk=project.pk)
        obj.save()
        log_activity(request.user, "questionnaires", "create", obj)
        return redirect("questionnaires:fill", pk=obj.pk)
    return render(request, "shared/form.html", {"form": form, "title": "Añadir ficha técnica", "subtitle": project.name})


@role_required("developer")
def fill(request, pk):
    questionnaire = get_object_or_404(
        ProjectQuestionnaire.objects
        .select_related(
            "project__client",
            "project__purchased_plan__plan",
            "template",
        )
        .prefetch_related(
            "template__sections__questions",
            "answers__question",
            "answers__updated_by",
        ),
        pk=pk,
    )

    if not can_access_project(
        request.user,
        questionnaire.project,
    ):
        raise PermissionDenied(
            "Este proyecto no está asignado a Desarrollo."
        )

    if questionnaire.template.code == WEBSITE_TEMPLATE_CODE:

        if request.method == "POST":

            section_focus = _clean(
                request.POST.get(
                    "section_focus"
                )
            )

            save_action = _clean(
                request.POST.get(
                    "save_action"
                )
            )

            allowed_sections = {
                "general",
                "business",
                "services",
                "coverage",
                "contact",
                "quotes",
                "brand",
                "certifications",
                "team",
                "marketing",
                "forms",
            }

            with transaction.atomic():

                progress = _website_save(
                    questionnaire,
                    request,
                )

                log_activity(
                    request.user,
                    "questionnaires",
                    "save_website_technical",
                    questionnaire,
                    description=(
                        f"Ficha técnica "
                        f"{progress['percent']}%"
                    ),
                )

            messages.success(
                request,
                (
                    "Ficha técnica guardada · "
                    f"{progress['percent']}% completado."
                ),
            )

            if (
                save_action == "keep"
                and section_focus in allowed_sections
            ):

                fill_url = reverse(
                    "questionnaires:fill",
                    kwargs={
                        "pk": pk
                    },
                )

                return redirect(
                    f"{fill_url}?section={section_focus}"
                )

            return redirect(
                "questionnaires:fill",
                pk=pk,
            )

        return render(
            request,
            "questionnaires/website_fill.html",
            _website_fill_context(
                questionnaire
            ),
        )

    existing = {
        a.question_id: a
        for a in questionnaire.answers.all()
    }

    sections = (
        questionnaire.template.sections
        .prefetch_related("questions")
        .all()
    )

    if request.method == "POST":

        with transaction.atomic():

            for section in sections:

                for question in section.questions.all():

                    field = f"q_{question.pk}"

                    state = request.POST.get(
                        f"q_state_{question.pk}",
                        AnswerState.PENDING,
                    )

                    if state not in AnswerState.values:
                        state = AnswerState.PENDING

                    if question.question_type == "multi":

                        value_list = request.POST.getlist(
                            field
                        )

                        Answer.objects.update_or_create(
                            questionnaire=questionnaire,
                            question=question,
                            defaults={
                                "state": state,
                                "value_text": "",
                                "value_json": {
                                    "values": value_list
                                },
                                "updated_by": request.user,
                            },
                        )

                    else:

                        value = request.POST.get(
                            field,
                            "",
                        )

                        Answer.objects.update_or_create(
                            questionnaire=questionnaire,
                            question=question,
                            defaults={
                                "state": state,
                                "value_text": value,
                                "value_json": {},
                                "updated_by": request.user,
                            },
                        )

            questionnaire.status = request.POST.get(
                "questionnaire_status",
                QuestionnaireStatus.IN_PROGRESS,
            )

            if (
                questionnaire.status
                not in QuestionnaireStatus.values
            ):
                questionnaire.status = (
                    QuestionnaireStatus.IN_PROGRESS
                )

            questionnaire.save(
                update_fields=[
                    "status",
                    "updated_at",
                ]
            )

            log_activity(
                request.user,
                "questionnaires",
                "save_answers",
                questionnaire,
            )

        messages.success(
            request,
            "Ficha técnica guardada.",
        )

        if (
            questionnaire.status
            == QuestionnaireStatus.COMPLETE
        ):
            return redirect(
                "questionnaires:development_dashboard"
            )

        return redirect(
            "questionnaires:fill",
            pk=pk,
        )

    answer_map = {}
    state_map = {}

    for qid, answer in existing.items():

        answer_map[qid] = (
            answer.value_json.get(
                "values",
                [],
            )
            if answer.value_json
            else answer.value_text
        )

        state_map[qid] = answer.state

    return render(
        request,
        "questionnaires/fill.html",
        {
            "questionnaire": questionnaire,
            "sections": sections,
            "answer_map": answer_map,
            "state_map": state_map,
            "answer_states": AnswerState.choices,
        },
    )